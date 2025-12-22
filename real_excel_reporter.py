"""
Real Excel Reporter - FULLY TRANSPARENT with ALL required fields
Includes: Data source proof, entry rules, AI confidence, daily P&L tracking
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Dict, List
import os


class RealExcelReporter:
    """
    Generates fully transparent Excel reports with ALL verification fields.
    """
    
    def __init__(self, output_folder: str):
        """
        Initialize Excel reporter.
        """
        self.output_folder = output_folder
        os.makedirs(output_folder, exist_ok=True)
    
    def generate_report(self, all_trades: List[Dict], filename: str):
        """
        Generate complete Excel report with 4 sheets.
        """
        filepath = os.path.join(self.output_folder, filename)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # SHEET 1: Daily Trade Log (FULLY TRANSPARENT)
            self._create_trade_log(all_trades, writer)
            
            # SHEET 2: Daily Result Summary (WITH CAPITAL TRACKING)
            self._create_daily_summary(all_trades, writer)
            
            # SHEET 3: Prompt Performance
            self._create_prompt_performance(all_trades, writer)
            
            # SHEET 4: Algorithm Config
            self._create_algorithm_config(writer)
        
        # Apply formatting
        self._apply_formatting(filepath)
        
        print(f"\n✅ Excel report generated: {filepath}")
    
    def _create_trade_log(self, trades: List[Dict], writer):
        """
        SHEET 1: Daily Trade Log - FULLY TRANSPARENT
        
        Includes ALL verification fields with clear column names.
        """
        if not trades:
            df = pd.DataFrame(columns=[
                'Date', 'PromptType', 'Stock', 'Bias',
                'DataSource', 'CandleTimeframe', 
                'EntryRule', 'EntryTime', 'EntryPrice',
                'StopLoss', 'Target',
                'ExitTime', 'ExitReason',
                'ExitPrice', 'Quantity', 'InvestedAmount',
                'AIConfidence', 'AIReason',
                'ProfitBeforeCosts', 'TransactionCost', 'FinalProfit', 'P&L%'
            ])
        else:
            # Import config for dynamic values
            from config_real import DATA_CONFIG, ALGORITHM_CONFIG
            
            # Create comprehensive trade log
            trade_data = []
            
            for trade in trades:
                # Get timeframe from trade (multi-timeframe) or config (single timeframe)
                timeframe = trade.get('timeframe', DATA_CONFIG.get('candle_interval', '5m'))
                
                # Calculate P&L percentage
                invested = trade.get('invested_amount', 0)
                final_pnl = trade.get('net_pnl', 0)
                pnl_percent = (final_pnl / invested * 100) if invested > 0 else 0
                
                # Get actual entry and exit times (not hardcoded 9:20)
                entry_time = trade.get('entry_time', 'N/A')
                exit_time = trade.get('exit_time', 'N/A')
                exit_reason = trade.get('exit_reason', '')
                
                # Make exit reason clearer
                if exit_reason == 'FORCE_EXIT_EOD':
                    exit_reason_display = 'Force Exit (End of Day)'
                elif exit_reason == 'STOP_LOSS':
                    exit_reason_display = 'Stop Loss Hit'
                elif exit_reason == 'TARGET':
                    exit_reason_display = 'Target Hit'
                else:
                    exit_reason_display = exit_reason
                
                trade_data.append({
                    'Date': trade.get('date', ''),
                    'PromptType': trade.get('prompt_type', ''),
                    'Stock': trade.get('symbol', ''),
                    'Bias': trade.get('bias', ''),
                    'DataSource': DATA_CONFIG['price_source'],
                    'CandleTimeframe': timeframe,
                    'EntryRule': f"First candle after {ALGORITHM_CONFIG['entry_start_time']}",
                    'EntryTime': entry_time,  # ACTUAL TIME
                    'EntryPrice': trade.get('entry_price', 0),
                    'StopLoss': trade.get('stop_loss', 0),
                    'Target': trade.get('target', 0),
                    'ExitTime': exit_time,  # ACTUAL TIME
                    'ExitReason': exit_reason_display,  # CLEAR REASON
                    'ExitPrice': trade.get('exit_price', 0),
                    'Quantity': trade.get('quantity', 0),
                    'InvestedAmount': trade.get('invested_amount', 0),
                    'AIConfidence': trade.get('confidence', 0),
                    'AIReason': trade.get('ai_reason', 'N/A'),
                    'ProfitBeforeCosts': trade.get('gross_pnl', 0),  # RENAMED
                    'TransactionCost': trade.get('transaction_cost', 0),
                    'FinalProfit': final_pnl,  # RENAMED
                    'P&L%': round(pnl_percent, 2)  # NEW COLUMN
                })
            
            df = pd.DataFrame(trade_data)
            
            # Sort by date and prompt type
            df = df.sort_values(['Date', 'PromptType', 'EntryTime'])
        
        df.to_excel(writer, sheet_name='Trade Log (Full Details)', index=False)
    
    def _create_daily_summary(self, trades: List[Dict], writer):
        """
        SHEET 2: Daily Result Summary - WITH CAPITAL TRACKING
        
        Shows:
        - CapitalStart (capital at start of day)
        - CapitalEnd (capital at end of day)
        - DailyP&L (profit/loss for the day)
        - DayStatus (PROFIT/LOSS/BREAKEVEN)
        """
        if not trades:
            df = pd.DataFrame(columns=[
                'Date', 'PromptType', 'TotalTrades',
                'CapitalStart', 'CapitalInvested', 'CapitalEnd',
                'Profit', 'Loss', 'DailyP&L', 'DayStatus'
            ])
        else:
            from config_real import ALGORITHM_CONFIG
            daily_capital = ALGORITHM_CONFIG['daily_capital']
            
            # Group by date and prompt type
            daily_results = []
            
            df_trades = pd.DataFrame(trades)
            
            for (date, prompt_type), group in df_trades.groupby(['date', 'prompt_type']):
                total_trades = len(group)
                capital_invested = group['invested_amount'].sum()
                profit = group[group['net_pnl'] > 0]['net_pnl'].sum()
                loss = abs(group[group['net_pnl'] < 0]['net_pnl'].sum())
                daily_pnl = group['net_pnl'].sum()
                
                # Capital tracking
                capital_start = daily_capital
                capital_end = capital_start + daily_pnl
                
                if daily_pnl > 0:
                    day_status = 'PROFIT'
                elif daily_pnl < 0:
                    day_status = 'LOSS'
                else:
                    day_status = 'BREAKEVEN'
                
                daily_results.append({
                    'Date': date,
                    'PromptType': prompt_type,
                    'TotalTrades': total_trades,
                    'CapitalStart': round(capital_start, 2),
                    'CapitalInvested': round(capital_invested, 2),
                    'CapitalEnd': round(capital_end, 2),
                    'Profit': round(profit, 2),
                    'Loss': round(loss, 2),
                    'DailyP&L': round(daily_pnl, 2),
                    'DayStatus': day_status
                })
            
            df = pd.DataFrame(daily_results)
            df = df.sort_values(['Date', 'PromptType'])
        
        df.to_excel(writer, sheet_name='Daily Summary', index=False)
    
    def _create_prompt_performance(self, trades: List[Dict], writer):
        """
        SHEET 3: Prompt Performance - WITH SIGNAL STRENGTH METRICS
        
        Includes:
        - Win Rate, Avg Win, Avg Loss (for expectancy calculation)
        - Expectancy (expected profit per trade)
        - Signal Strength (normalized expectancy for position sizing)
        """
        if not trades:
            df = pd.DataFrame(columns=[
                'PromptType', 'TotalTrades', 'WinRate%',
                'TotalDays', 'ProfitableDays', 'LosingDays',
                'AvgWin', 'AvgLoss', 'Expectancy', 'SignalStrength',
                'TotalInvested', 'TotalProfit', 'TotalLoss', 'NetProfit/Loss'
            ])
        else:
            df_trades = pd.DataFrame(trades)
            
            # Group by prompt type
            performance = []
            
            for prompt_type, group in df_trades.groupby('prompt_type'):
                # Group by date to count days
                daily_pnl = group.groupby('date')['net_pnl'].sum()
                
                total_trades = len(group)
                total_days = len(daily_pnl)
                profitable_days = (daily_pnl > 0).sum()
                losing_days = (daily_pnl < 0).sum()
                
                # Calculate win rate
                winning_trades = (group['net_pnl'] > 0).sum()
                losing_trades = (group['net_pnl'] < 0).sum()
                win_rate = (winning_trades / total_trades * 100) if total_trades > 0 else 0
                
                # Calculate average win and average loss
                wins = group[group['net_pnl'] > 0]['net_pnl']
                losses = group[group['net_pnl'] < 0]['net_pnl']
                
                avg_win = wins.mean() if len(wins) > 0 else 0
                avg_loss = abs(losses.mean()) if len(losses) > 0 else 0
                
                # Calculate expectancy: (WinRate × AvgWin) - (LossRate × AvgLoss)
                loss_rate = (losing_trades / total_trades) if total_trades > 0 else 0
                expectancy = (win_rate/100 * avg_win) - (loss_rate * avg_loss)
                
                # Signal strength (clipped to 0 if negative)
                signal_strength = max(0, expectancy)
                
                # Calculate invested, profit, loss
                total_invested = group['invested_amount'].sum()
                total_profit = group[group['net_pnl'] > 0]['net_pnl'].sum()
                total_loss = abs(group[group['net_pnl'] < 0]['net_pnl'].sum())
                net_outcome = group['net_pnl'].sum()
                
                performance.append({
                    'PromptType': prompt_type,
                    'TotalTrades': total_trades,
                    'WinRate%': round(win_rate, 1),
                    'TotalDays': total_days,
                    'ProfitableDays': profitable_days,
                    'LosingDays': losing_days,
                    'AvgWin': round(avg_win, 2),
                    'AvgLoss': round(avg_loss, 2),
                    'Expectancy': round(expectancy, 2),
                    'SignalStrength': round(signal_strength, 2),
                    'TotalInvested': round(total_invested, 2),
                    'TotalProfit': round(total_profit, 2),
                    'TotalLoss': round(total_loss, 2),
                    'NetProfit/Loss': round(net_outcome, 2)
                })
            
            df = pd.DataFrame(performance)
            df = df.sort_values('SignalStrength', ascending=False)  # Sort by signal strength
            
            # Add OVERALL SUMMARY row
            overall_wins = df_trades[df_trades['net_pnl'] > 0]['net_pnl']
            overall_losses = df_trades[df_trades['net_pnl'] < 0]['net_pnl']
            overall_winning_trades = (df_trades['net_pnl'] > 0).sum()
            overall_losing_trades = (df_trades['net_pnl'] < 0).sum()
            overall_total_trades = len(df_trades)
            
            overall_win_rate = (overall_winning_trades / overall_total_trades * 100) if overall_total_trades > 0 else 0
            overall_avg_win = overall_wins.mean() if len(overall_wins) > 0 else 0
            overall_avg_loss = abs(overall_losses.mean()) if len(overall_losses) > 0 else 0
            overall_loss_rate = (overall_losing_trades / overall_total_trades) if overall_total_trades > 0 else 0
            overall_expectancy = (overall_win_rate/100 * overall_avg_win) - (overall_loss_rate * overall_avg_loss)
            overall_signal_strength = max(0, overall_expectancy)
            
            overall_row = {
                'PromptType': '═══ OVERALL (All Prompts) ═══',
                'TotalTrades': overall_total_trades,
                'WinRate%': round(overall_win_rate, 1),
                'TotalDays': df_trades['date'].nunique(),
                'ProfitableDays': len(df_trades.groupby('date')['net_pnl'].sum()[df_trades.groupby('date')['net_pnl'].sum() > 0]),
                'LosingDays': len(df_trades.groupby('date')['net_pnl'].sum()[df_trades.groupby('date')['net_pnl'].sum() < 0]),
                'AvgWin': round(overall_avg_win, 2),
                'AvgLoss': round(overall_avg_loss, 2),
                'Expectancy': round(overall_expectancy, 2),
                'SignalStrength': round(overall_signal_strength, 2),
                'TotalInvested': round(df_trades['invested_amount'].sum(), 2),
                'TotalProfit': round(df_trades[df_trades['net_pnl'] > 0]['net_pnl'].sum(), 2),
                'TotalLoss': round(abs(df_trades[df_trades['net_pnl'] < 0]['net_pnl'].sum()), 2),
                'NetProfit/Loss': round(df_trades['net_pnl'].sum(), 2)
            }
            
            # Append overall row
            df = pd.concat([df, pd.DataFrame([overall_row])], ignore_index=True)
        
        df.to_excel(writer, sheet_name='Prompt Performance', index=False)
    
    def _create_algorithm_config(self, writer):
        """
        SHEET 4: Algorithm Config (Transparency)
        """
        from config_real import ALGORITHM_CONFIG, DATA_CONFIG
        
        config_data = {
            'Parameter': [
                'StopLossPercent',
                'TargetPercent',
                'RiskReward',
                'CapitalPerDay',
                'CapitalPerTrade',
                'MaxPositions',
                'CandleTimeframe',
                'DataSource',
                'NewsSource',
                'AIModel',
                'EntryStartTime',
                'ForceExitTime',
                'SlippagePercent',
                'TransactionCostPercent'
            ],
            'Value': [
                f"{ALGORITHM_CONFIG['stop_loss_percent']}%",
                f"{ALGORITHM_CONFIG['target_percent']}%",
                f"1:{ALGORITHM_CONFIG['risk_reward_ratio']}",
                f"₹{ALGORITHM_CONFIG['daily_capital']:,}",
                f"₹{ALGORITHM_CONFIG['capital_per_trade']:,}",
                ALGORITHM_CONFIG['max_positions_per_day'],
                DATA_CONFIG.get('candle_interval', ', '.join(DATA_CONFIG.get('candle_intervals_to_test', ['5m']))),
                DATA_CONFIG['price_source'],
                DATA_CONFIG['news_source'],
                DATA_CONFIG['ai_model'],
                ALGORITHM_CONFIG['entry_start_time'],
                ALGORITHM_CONFIG['force_exit_time'],
                f"{ALGORITHM_CONFIG['slippage_percent']}%",
                f"{ALGORITHM_CONFIG['transaction_cost_percent']}%"
            ]
        }
        
        df = pd.DataFrame(config_data)
        df.to_excel(writer, sheet_name='Algorithm Config', index=False)
    
    def _apply_formatting(self, filepath: str):
        """
        Apply professional formatting to Excel file with enhanced colors.
        """
        wb = load_workbook(filepath)
        
        # Header style
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        
        # Profit/Loss colors (more vibrant)
        profit_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        profit_font = Font(bold=True, color='FFFFFF')
        
        loss_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        loss_font = Font(bold=True, color='FFFFFF')
        
        # Alternating row colors
        alt_row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # Overall summary row style
        overall_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
        overall_font = Font(bold=True, size=11)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Alternating row colors
            for row in range(2, ws.max_row + 1):
                if row % 2 == 0:
                    for cell in ws[row]:
                        if cell.fill.start_color.index == '00000000':
                            cell.fill = alt_row_fill
            
            # Conditional formatting for P&L
            if sheet_name == 'Trade Log (Full Details)':
                # FinalProfit is column U (21st column), P&L% is column V (22nd column)
                for row in range(2, ws.max_row + 1):
                    # Color FinalProfit
                    final_profit_cell = ws[f'U{row}']
                    try:
                        profit_value = float(final_profit_cell.value)
                        if profit_value > 0:
                            final_profit_cell.fill = profit_fill
                            final_profit_cell.font = profit_font
                        elif profit_value < 0:
                            final_profit_cell.fill = loss_fill
                            final_profit_cell.font = loss_font
                    except:
                        pass
                    
                    # Color P&L%
                    pnl_percent_cell = ws[f'V{row}']
                    try:
                        percent_value = float(pnl_percent_cell.value)
                        if percent_value > 0:
                            pnl_percent_cell.fill = profit_fill
                            pnl_percent_cell.font = profit_font
                        elif percent_value < 0:
                            pnl_percent_cell.fill = loss_fill
                            pnl_percent_cell.font = loss_font
                    except:
                        pass
                    
                    # Also color ProfitBeforeCosts (column S)
                    gross_pnl_cell = ws[f'S{row}']
                    try:
                        gross_value = float(gross_pnl_cell.value)
                        if gross_value > 0:
                            gross_pnl_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        elif gross_value < 0:
                            gross_pnl_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    except:
                        pass
            
            elif sheet_name == 'Daily Summary':
                # DailyP&L is column I, DayStatus is column J
                for row in range(2, ws.max_row + 1):
                    # Color DailyP&L
                    pnl_cell = ws[f'I{row}']
                    try:
                        pnl_value = float(pnl_cell.value)
                        if pnl_value > 0:
                            pnl_cell.fill = profit_fill
                            pnl_cell.font = profit_font
                        elif pnl_value < 0:
                            pnl_cell.fill = loss_fill
                            pnl_cell.font = loss_font
                    except:
                        pass
                    
                    # Color DayStatus
                    status_cell = ws[f'J{row}']
                    if status_cell.value == 'PROFIT':
                        status_cell.fill = profit_fill
                        status_cell.font = profit_font
                    elif status_cell.value == 'LOSS':
                        status_cell.fill = loss_fill
                        status_cell.font = loss_font
            
            elif sheet_name == 'Prompt Performance':
                # Expectancy is column I, SignalStrength is column J, NetProfit/Loss is column N
                for row in range(2, ws.max_row + 1):
                    # Check if this is the OVERALL row
                    prompt_cell = ws[f'A{row}']
                    is_overall = '═══ OVERALL' in str(prompt_cell.value)
                    
                    if is_overall:
                        # Special formatting for OVERALL row
                        for cell in ws[row]:
                            cell.fill = overall_fill
                            cell.font = overall_font
                    
                    # Color Expectancy
                    expectancy_cell = ws[f'I{row}']
                    try:
                        expectancy_value = float(expectancy_cell.value)
                        if expectancy_value > 0:
                            expectancy_cell.fill = profit_fill
                            expectancy_cell.font = profit_font
                        elif expectancy_value < 0:
                            expectancy_cell.fill = loss_fill
                            expectancy_cell.font = loss_font
                    except:
                        pass
                    
                    # Color SignalStrength
                    signal_cell = ws[f'J{row}']
                    try:
                        signal_value = float(signal_cell.value)
                        if signal_value > 0:
                            signal_cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                            signal_cell.font = Font(bold=True)
                    except:
                        pass
                    
                    # Color NetProfit/Loss
                    outcome_cell = ws[f'N{row}']
                    try:
                        outcome_value = float(outcome_cell.value)
                        if outcome_value > 0:
                            outcome_cell.fill = profit_fill
                            outcome_cell.font = profit_font
                        elif outcome_value < 0:
                            outcome_cell.fill = loss_fill
                            outcome_cell.font = loss_font
                    except:
                        pass
        
        wb.save(filepath)
